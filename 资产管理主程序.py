import os
import csv
import json
import time
import sys
import random

# 依赖检查
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
except ImportError:
    print(" 缺少 openpyxl，请运行 pip install openpyxl")
    sys.exit(1)

# === 配置区 ===
TARGET_ROOT = os.getcwd()
CONFIG_FILE = "00_系统配置_俏皮话.csv"
HISTORY_FILE = "system_history.json"
OUTPUT_EXCEL = "融资资产全景导航表.xlsx"
IGNORE_DIRS = {'logs', 'dist', 'build', '.git', '__pycache__', 'venv', '交付给团队_最终版', '交付_新人标准开荒包_v1.0'}

# 默认语录 (兜底用)
QUOTES = {
    "LOADING": ["系统启动中..."],
    "SCAN_START": ["开始扫描..."],
    "GOOD_INDEX": ["不错！"],
    "BAD_NAKED": ["发现问题！"],
    "SCORE_UP": ["有进步！"],
    "SCORE_DOWN": ["退步了..."],
    "GENESIS_SET": ["基准线已设置。"],
    "GENESIS_EXIST": ["基准线存在。"]
}

# === 情感引擎 ===
def load_quotes():
    """加载俏皮话配置"""
    global QUOTES
    if not os.path.exists(CONFIG_FILE): return
    try:
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                t = row.get("类型")
                msg = row.get("语录内容")
                if t and msg:
                    if t not in QUOTES: QUOTES[t] = []
                    QUOTES[t].append(msg)
    except: pass

def say(type_key, extra=""):
    """随机说一句骚话"""
    msgs = QUOTES.get(type_key, ["..."])
    msg = random.choice(msgs)
    print(f" {msg} {extra}")
    time.sleep(0.3) # 稍微停顿，让人看清

def loading_animation():
    """假装很忙的进度条"""
    print("")
    msgs = QUOTES.get("LOADING", ["加载中..."])
    total = 20
    for i in range(total + 1):
        time.sleep(0.05 + random.random() * 0.05) # 随机卡顿
        percent = int((i / total) * 100)
        bar = "" * i + "" * (total - i)
        
        # 随机挑一句骚话显示在进度条后面
        current_msg = msgs[i % len(msgs)] if msgs else "..."
        sys.stdout.write(f"\r [{bar}] {percent}% | {current_msg}")
        sys.stdout.flush()
    print("\n")

# === 核心逻辑 ===
def parse_index(filepath):
    """解析索引质量"""
    info = {"is_valid": False, "fields": [], "missing": []}
    required = ["名称", "版本", "负责人", "日期"] # 简化关键词
    
    try:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()
        
        info["is_valid"] = True
        content_str = "".join(lines[:30]) # 只读前30行
        
        for req in required:
            if req in content_str:
                info["fields"].append(req)
            else:
                info["missing"].append(req)
    except: pass
    return info

def run_main():
    load_quotes()
    loading_animation() # 播放开场动画
    
    say("SCAN_START")
    
    scan_data = []
    valid_count = 0
    naked_count = 0
    
    # === 1. 全盘扫描 ===
    for root, dirs, files in os.walk(TARGET_ROOT):
        dirs[:] = [d for d in dirs if d not in IGNORE_DIRS]
        if root == TARGET_ROOT: continue
        
        file_list = [f for f in files if not f.startswith("~")]
        if not file_list: continue
        
        rel_path = root.replace(TARGET_ROOT, "")
        folder_item = {"path": rel_path, "status": "unknown", "missing": []}
        
        index_files = [f for f in file_list if f.startswith("00_") and f.endswith(".txt")]
        
        if index_files:
            folder_item["status"] = "valid"
            info = parse_index(os.path.join(root, index_files[0]))
            folder_item["missing"] = info["missing"]
            valid_count += 1
            # 随机表扬
            if random.random() > 0.8: say("GOOD_INDEX", f"({rel_path})")
        else:
            folder_item["status"] = "naked"
            naked_count += 1
            # 随机吐槽
            if random.random() > 0.8: say("BAD_NAKED", f"({rel_path})")
            
        scan_data.append(folder_item)
        
    # === 2. 算分 (简单粗暴版) ===
    # 健康分：100 起扣
    health_score = 100 - (naked_count * 10)
    for item in scan_data:
        health_score -= (len(item["missing"]) * 2)
    health_score = max(0, health_score)

    # === 3. 历史基准线 (双轨制) ===
    history = {}
    if os.path.exists(HISTORY_FILE):
        try:
            with open(HISTORY_FILE, 'r') as f: history = json.load(f)
        except: pass
        
    # 读取/设置 创世纪基准
    genesis = history.get("genesis")
    is_genesis_now = False
    
    if not genesis:
        print("\n" + "="*40)
        print(" 战友，这是我们第一次见面（或者你删了存档）。")
        ans = input(" 是否将此刻定格为【创世纪元年】基准线？(输入 Y 确认，直接回车跳过): ").strip().upper()
        if ans == "Y":
            genesis = {"date": time.strftime('%Y-%m-%d'), "valid": valid_count, "score": health_score}
            history["genesis"] = genesis
            is_genesis_now = True
            say("GENESIS_SET")
    else:
        say("GENESIS_EXIST", f"({genesis['date']} 建立)")

    # 计算增量
    growth_count = 0
    if genesis:
        growth_count = valid_count - genesis["valid"]
    
    # 记录本次运行
    run_record = {
        "date": time.strftime('%Y-%m-%d %H:%M'),
        "health": health_score,
        "valid": valid_count,
        "growth": growth_count
    }
    if "logs" not in history: history["logs"] = []
    history["logs"].append(run_record)
    
    # 保存历史
    with open(HISTORY_FILE, 'w') as f: json.dump(history, f, indent=2)

    # === 4. 生成 Excel (三维报表) ===
    wb = Workbook()
    
    # Sheet 1: 指挥官驾驶舱
    ws1 = wb.active
    ws1.title = "指挥官驾驶舱"
    ws1.merge_cells('A1:D2')
    ws1['A1'] = f"存量健康分: {health_score}"
    ws1['A1'].font = Font(size=30, bold=True, color="008000" if health_score >= 80 else "FF0000")
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws1.merge_cells('E1:H2')
    growth_str = f"+{growth_count}" if growth_count > 0 else str(growth_count)
    ws1['E1'] = f"创世纪以来增量: {growth_str} 项"
    ws1['E1'].font = Font(size=20, bold=True, color="0000FF")
    ws1['E1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Sheet 2: 资产明细
    ws2 = wb.create_sheet("资产明细")
    ws2.append(["物理路径", "状态", "缺失字段", "整改建议"])
    for item in scan_data:
        status_text = " 合规" if item["status"] == "valid" else " 裸奔"
        missing_text = ",".join(item["missing"]) if item["missing"] else "无"
        advice = "补写索引" if item["status"] == "naked" else ("补全字段" if item["missing"] else "完美")
        
        ws2.append([item["path"], status_text, missing_text, advice])
        
        # 上色
        if item["status"] == "naked":
            for cell in ws2[ws2.max_row]: cell.fill = PatternFill(start_color="FFCCCC", fill_type="solid")
            
    # Sheet 3: 历史军功簿
    ws3 = wb.create_sheet("历史军功簿")
    ws3.append(["扫描时间", "健康分", "合规资产数", "较创世纪增量"])
    for log in history["logs"]:
        ws3.append([log["date"], log["health"], log["valid"], log["growth"]])
        
    wb.save(OUTPUT_EXCEL)
    
    print("\n" + "="*40)
    print(f" 报告已生成: {OUTPUT_EXCEL}")
    
    # 最终评价
    if growth_count > 0: say("SCORE_UP")
    elif health_score < 60: say("SCORE_DOWN")
    
    print("\n 按回车键退出战场... (别忘了给战友倒杯水)")
    input()

if __name__ == "__main__":
    run_main()